import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from config import DB_FILE, ATTENDANCE_FILE, EMPLOYEE_SHEET, ATTENDANCE_SHEET
from config import DB_FILE

class ExcelDB:

    def __init__(self):
        self._init_employees_db()
        self._init_attendance_db()

    # ── Employees DB ──────────────────
    def _init_employees_db(self):
        if os.path.exists(DB_FILE):
            wb = load_workbook(DB_FILE)

            if EMPLOYEE_SHEET not in wb.sheetnames:
                ws = wb.active
                ws.title = EMPLOYEE_SHEET

                headers = ["Employee ID", "Full Name", "Department", "Position"]
                self._write_header(ws, headers)

                wb.save(DB_FILE)
            return

        # Create new file
        wb = Workbook()
        ws = wb.active
        ws.title = EMPLOYEE_SHEET

        headers = ["Employee ID", "Full Name", "Department", "Position"]
        self._write_header(ws, headers)

        seed = [
            ("E001", "Alice Johnson", "Engineering", "Senior Developer"),
            ("E002", "Bob Smith", "Marketing", "Marketing Manager"),
            ("E003", "Carol Williams", "HR", "HR Specialist"),
        ]

        for row in seed:
            ws.append(row)

        wb.save(DB_FILE)

    # ── Attendance DB ──────────────────
    def _init_attendance_db(self):
        if os.path.exists(ATTENDANCE_FILE):
            wb = load_workbook(ATTENDANCE_FILE)

            if ATTENDANCE_SHEET not in wb.sheetnames:
                ws = wb.active
                ws.title = ATTENDANCE_SHEET

                headers = ["Record ID", "Employee ID", "Full Name",
                           "Department", "Date", "Time", "Status", "Marked By"]

                self._write_header(ws, headers)
                wb.save(ATTENDANCE_FILE)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = ATTENDANCE_SHEET

        headers = ["Record ID", "Employee ID", "Full Name",
                   "Department", "Date", "Time", "Status", "Marked By"]

        self._write_header(ws, headers)
        wb.save(ATTENDANCE_FILE)

    # ── Public API ──────────────────
    def get_employee(self, emp_id: str):
        try:
            wb = load_workbook(DB_FILE, read_only=True)
            ws = wb[EMPLOYEE_SHEET]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).upper() == emp_id.upper():
                    return {
                        "id": row[0],
                        "name": row[1],
                        "department": row[2],
                        "position": row[3],
                    }
        except Exception as e:
            raise RuntimeError(f"Error reading employee DB: {e}")

        return None

    def save_attendance(self, emp, status):
        try:
            wb = load_workbook(ATTENDANCE_FILE)
            ws = wb[ATTENDANCE_SHEET]

            record_id = ws.max_row
            now = datetime.now()

            ws.append([
                record_id,
                emp["id"],
                emp["name"],
                emp["department"],
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                status,
                "System"
            ])

            wb.save(ATTENDANCE_FILE)
            return record_id

        except Exception as e:
            raise RuntimeError(f"Error saving attendance: {e}")

    # ── Helpers ──────────────────
    @staticmethod
    def _write_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)